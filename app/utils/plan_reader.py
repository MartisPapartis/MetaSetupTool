"""Plan file reader — extracts Meta campaign rows from the "Planas" sheet.

Sheet: "Planas"
Column mapping:
  AD  → filter: only rows where value == "Meta" (case-insensitive)
  AK  → start date
  AL  → end date
  AP  → lifetime budget amount (EUR); if populated use lifetime_budget, else skip budget
  BN  → campaign name
  BO  → ad set name
"""

from __future__ import annotations

import datetime

from app.models.campaign_data import CampaignData, AdSetData, AdData, CreativeData, default_ad_name
from app.models.enums import AdFormat, AdObjective, AdStatus, OptimizationGoal, BillingEvent


def _col_index(letter: str) -> int:
    """Convert Excel column letter(s) to 0-based index."""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


# 0-based column indices
_COL_AD = _col_index("AD")
_COL_AK = _col_index("AK")
_COL_AL = _col_index("AL")
_COL_AP = _col_index("AP")
_COL_BN = _col_index("BN")
_COL_BO = _col_index("BO")


def _get(row: tuple, idx: int):
    try:
        return row[idx]
    except IndexError:
        return None


def _str(row: tuple, idx: int) -> str:
    val = _get(row, idx)
    return str(val).strip() if val is not None else ""


def _date_only(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip().split(" ", maxsplit=1)[0].split("T", maxsplit=1)[0]



def _process_plan_row(row: tuple, campaigns: dict[str, CampaignData]) -> None:
    if _str(row, _COL_AD).lower() != "meta":
        return

    campaign_name = _str(row, _COL_BN)
    if not campaign_name:
        return

    adset_name = _str(row, _COL_BO)
    if not adset_name:
        return

    if campaign_name not in campaigns:
        budget_raw = _get(row, _COL_AP)
        try:
            has_budget = budget_raw not in (None, "")
            budget_cents = int(float(budget_raw) * 100) if has_budget else None
        except (ValueError, TypeError):
            budget_cents = None
        campaigns[campaign_name] = CampaignData(
            name=campaign_name,
            objective=AdObjective.OUTCOME_TRAFFIC,
            status=AdStatus.PAUSED,
            lifetime_budget=budget_cents,
        )
    campaign = campaigns[campaign_name]

    adset = next((a for a in campaign.ad_sets if a.name == adset_name), None)
    if adset is None:
        start_date = _date_only(_get(row, _COL_AK))
        end_date = _date_only(_get(row, _COL_AL))
        adset = AdSetData(
            name=adset_name,
            status=AdStatus.PAUSED,
            optimization_goal=OptimizationGoal.LINK_CLICKS,
            billing_event=BillingEvent.IMPRESSIONS,
            start_time=f"{start_date} 00:00:00" if start_date else "",
            end_time=f"{end_date} 23:59:59" if end_date else "",
        )
        campaign.ad_sets.append(adset)

    creative = CreativeData()
    adset.ads.append(AdData(
        name=default_ad_name(AdFormat.SINGLE_IMAGE),
        status=AdStatus.PAUSED,
        creative=creative,
    ))


def read_campaigns_from_plan(path: str) -> list[CampaignData]:
    """Parse a Plan .xlsx file and return CampaignData objects."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True)

    if "Planas" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Planas' not found in {path}. Available: {wb.sheetnames}")

    ws = wb["Planas"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    campaigns: dict[str, CampaignData] = {}

    for row in rows[1:]:
        if not any(row):
            continue
        _process_plan_row(row, campaigns)

    return list(campaigns.values())
